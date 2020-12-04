import openpyxl


def get_column_and_row(team):
    if team == "Hendrix":
        return "B", "2"
    if team == "Birmingham-Southern":
        return "C", "3"
    if team == "Millsaps":
        return "D", "4"
    if team == "Centre":
        return "E", "5"
    if team == "Sewanee":
        return "F", "6"
    if team == "Berry":
        return "G", "7"
    if team == "Oglethorpe":
        return "H", "8"
    if team == "Rhodes":
        return "I", "9"
    else:
        return "J", "10"


def get_reference(team1, team2):
    team1_col, team1_row = get_column_and_row(team1)
    team2_col, team2_row = get_column_and_row(team2)
    team1_ref = team1_col + team2_row
    team2_ref = team2_col + team1_row
    return team1_ref, team2_ref


def get_match_data(ws, end_of_worksheet):
    matches = []
    for i in range(1, end_of_worksheet):
        if ws[f'E{i}'].value == '(Conf.)':
            team1 = ws[f'B{i}'].value
            team2 = ws[f'C{i}'].value
            match_data = extract_match_data(team1, team2)
            if match_data is not None:
                matches.append(match_data)
    return matches


def extract_match_data(team1, team2):
    team1 = team1.split()
    team2 = team2.split()
    if len(team1) > 1 and len(team2) > 1 and team1[0] != team2[0]:
        match_data = {}
        team1_score = int(team1[1])
        team2_score = int(team2[1])
        match_data['team1'] = team1[0]
        match_data['team1_score'] = team1_score
        match_data['team2'] = team2[0]
        match_data['team2_score'] = team2_score
        return match_data
    return None


def write_data(ws, matches):
    for match in matches:
        team_1_cell, team_2_cell = get_reference(match['team1'], match['team2'])
        if ws[team_1_cell].value is None:
            ws[team_1_cell] = 0
        if ws[team_2_cell].value is None:
            ws[team_2_cell] = 0
        ws[team_1_cell] = ws[team_1_cell].value + match["team1_score"]
        ws[team_2_cell] = ws[team_2_cell].value + match["team2_score"]


if __name__ == '__main__':
    def main():
        file_name = "C:\\Users\\sammy\\Documents\\Classes\\2020-2021\\Fall Semester\\LinearAlgebra\\AppliedProject\\WomensLacrosse.xlsx"
        wb = openpyxl.load_workbook(filename=file_name)
        for year in range(2013, 2020):
            data_sheet = wb[f"{year}Data"]
            stat_sheet = wb[f"{year} Stats"]
            matches = get_match_data(data_sheet, 500)
            print(matches)
            write_data(stat_sheet, matches)
        wb.save(file_name)


    main()
